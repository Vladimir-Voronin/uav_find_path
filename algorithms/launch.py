import math
import time
import tracemalloc

from qgis.core import *
from ModuleInstruments.Converter import Converter
from ModuleInstruments.DebugLog import DebugLog
from ModuleInstruments.FindPathData import FindPathData
from algorithms.APFMethod import APFMethod
from algorithms.GdalUAV.processing.calculations.ObjectsCalculations import length_of_path_from_feats_lines
import csv
import openpyxl

from algorithms.GdalUAV.exceptions.MethodsException import TimeToSucceedException, FailFindPathException
from pympler import muppy

all_objects = muppy.get_objects()


class Couple:
    def __init__(self, point_start, point_target):
        self.point_start = point_start
        self.point_target = point_target
        self.base_length = self.__get_length()

    def __get_length(self):
        point_1 = self.point_start.asPoint()
        point_2 = self.point_target.asPoint()
        x_full_difference = point_2.x() - point_1.x()
        y_full_difference = point_2.y() - point_1.y()
        result = math.sqrt(x_full_difference ** 2 + y_full_difference ** 2)
        return (result * result) ** 0.5


class CsvReader:
    @staticmethod
    def get_couple_of_points(file):
        list_of_couples = []
        with open(file, mode='r', newline='') as csv_file:
            csv_reader = csv.reader(csv_file)
            line_count = 0
            for row in csv_reader:
                a = ", ".join(row)
                splited = a.split(';')
                res = []
                for i in splited:
                    res.append(i.split(','))

                for i in range(len(res)):
                    for k in range(len(res[i])):
                        if res[i][k] != "":
                            res[i][k] = float(res[i][k])

                list_of_couples.append([QgsGeometry.fromPointXY(QgsPointXY(res[0][0], res[0][1])),
                                        QgsGeometry.fromPointXY(QgsPointXY(res[1][0], res[1][1]))])
                line_count += 1
            print(f'Processed {line_count} lines.')

        result_list = []
        for i in list_of_couples:
            result_list.append(Couple(i[0], i[1]))

        return result_list


class Test:
    @staticmethod
    def run_test(list_of_couples, methods_list):
        QgsApplication.setPrefixPath(r'C:\OSGEO4~1\apps\qgis', True)
        qgs = QgsApplication([], False)
        qgs.initQgis()

        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(r"C:\Users\Neptune\Desktop\results.xlsx")
        try:
            # Define variable to read the active sheet:
            worksheet = wookbook.active
            current_row = worksheet.max_row + 1
            for method in methods_list:
                point_number = 0
                for points in list_of_couples:
                    print(f"start: {points.point_start}")
                    print(f"target: {points.point_target}")
                    print(f"point_numb: {point_number}")
                    point_number += 1
                    if point_number < 0:
                        continue
                    n = 1
                    length_full = 0
                    full_time = 0
                    number_of_obstacles = 1
                    result = True
                    full_memory = 0
                    area_precent = 0
                    for i in range(n):
                        proj = QgsProject.instance()
                        proj.read(r'C:\Users\Neptune\Desktop\Voronin qgis\Voronin qgis.qgs')
                        point1 = points.point_start
                        point2 = points.point_target
                        path = r"C:\Users\Neptune\Desktop\Voronin qgis\shp\Строения.shp"

                        obstacles = QgsVectorLayer(path)
                        source_list_of_geometry_obstacles = Converter.get_list_of_poligons_in_3395(obstacles, proj)
                        find_path_data = FindPathData(proj, point1, point2, obstacles,
                                                      r"C:\Users\Neptune\Desktop\Voronin qgis\shp",
                                                      False,
                                                      source_list_of_geometry_obstacles)
                        debug_log = DebugLog()
                        my_time = time.perf_counter()
                        check = method(find_path_data, debug_log)
                        try:
                            # starting the monitoring
                            tracemalloc.start()
                            _, start_peak = tracemalloc.get_traced_memory()
                            check.run()
                            _, final_peak = tracemalloc.get_traced_memory()
                            tracemalloc.stop()
                            full_memory = final_peak - start_peak
                            # area_precent = check.get_area_precents()

                        except TimeToSucceedException as t:
                            print(t.txt)
                            print("")
                            result = False
                        except FailFindPathException as t:
                            print(t.txt)
                            print("Fail")
                            result = False

                        number_of_obstacles = len(check.list_of_obstacles_geometry)
                        my_time = time.perf_counter() - my_time
                        full_time += my_time
                        if not check.final_path:
                            result = False
                        else:
                            length_full += length_of_path_from_feats_lines(check.final_path)

                    if not result:
                        length_full = 0
                        full_time = 0
                        number_of_obstacles = 0
                    full_time /= n
                    full_memory /= n
                    number_of_obstacles /= n
                    print(method.__name__)
                    print(full_time)
                    print(f"{full_memory} b")
                    print(f"n obst: {number_of_obstacles}")
                    length_full /= n
                    print(f"Length: {length_full}")
                    worksheet.cell(current_row, 1, value=points.base_length)
                    worksheet.cell(current_row, 2, value=method.__name__)
                    worksheet.cell(current_row, 3, value=full_time)
                    worksheet.cell(current_row, 4, value=length_full)
                    worksheet.cell(current_row, 5, value=n)
                    worksheet.cell(current_row, 6, value=full_memory)
                    worksheet.cell(current_row, 7, value=number_of_obstacles)
                    # worksheet.cell(current_row, 8, value=area_precent)

                    current_row += 1
                    wookbook.save(r"C:\Users\Neptune\Desktop\results.xlsx")
                    if full_time > 90:
                        break
        finally:
            wookbook.close()


if __name__ == '__main__':
    list_of_couples = CsvReader.get_couple_of_points(r"C:\Users\Neptune\Desktop\points_auto.csv")
    # method_list = [AStarMethod, FormerMethod, AStarMethodGrid, DijkstraMethod, BugMethod, RandomizedRoadmapGridMethod,
    #                RRTDirectMethod, DStarMethod, APFMethodOptimize, APFMethod]
    method_list = [APFMethod]
    Test.run_test(list_of_couples, method_list)
    pass
